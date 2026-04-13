#!/usr/bin/env python3
"""
ADF Trigger Performance Comparison Report
==========================================
Reads two raw data sheets from an Excel workbook and generates a formatted
"Comparison" sheet comparing ADF trigger run performance between a Legacy
environment and a Migrated environment.

The comparison evaluates:
  - Run counts for succeeded executions in each environment
  - Average duration (seconds) per trigger in each environment
  - Verdict: Faster / Slower / Comparable / Insufficient Runs / No Data
  - Difference in minutes between average durations

Input workbook (ADF_Trigger_Raw.xlsx) must contain two sheets:
  - Legacy_Raw   : trigger run data from the legacy environment
  - Migrated_Raw : trigger run data from the migrated environment

Expected columns in each raw sheet (1-indexed):
  A: Trigger Name
  B: Pipeline Name
  E: Run Status  (e.g. "Succeeded")
  F: Duration    (seconds, numeric)

Output:
  Overwrites the "Comparison" sheet in the same workbook with a
  colour-coded, formatted comparison table.

Usage:
  python adf_trigger_comparison.py
  python adf_trigger_comparison.py --input my_workbook.xlsx
  python adf_trigger_comparison.py --input my_workbook.xlsx --min-runs 3 --threshold 0.15

Verdict logic:
  - Migrated has 0 runs           → "No Migrated Data"
  - Migrated has < MIN_RUNS runs  → "Insufficient Runs"
  - Legacy has 0 runs             → "No Legacy Baseline"
  - Migrated avg ≤ Legacy avg × (1 - THRESHOLD) → "Faster"
  - Migrated avg ≥ Legacy avg × (1 + THRESHOLD) → "Slower"
  - Otherwise                     → "Comparable"
"""

import argparse
import sys

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# Constants — sheet names and comparison thresholds
# ─────────────────────────────────────────────────────────────────────────────

LEGACY_SHEET   = "Legacy_Raw"
MIGRATED_SHEET = "Migrated_Raw"
OUTPUT_SHEET   = "Comparison"

DEFAULT_MIN_RUNS   = 5     # Minimum migrated runs required for a valid verdict
DEFAULT_THRESHOLD  = 0.10  # 10% difference threshold for Faster/Slower verdict


# ─────────────────────────────────────────────────────────────────────────────
# Styles
# ─────────────────────────────────────────────────────────────────────────────

def _thin_border() -> Border:
    side = Side(border_style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


# Header row — dark blue background, white bold text
HEADER_FILL      = PatternFill("solid", start_color="1F4E78")
HEADER_FONT      = Font(name="Arial", bold=True, color="FFFFFF", size=11)

# Sub-header row — light blue background, dark blue bold text
SUB_FILL         = PatternFill("solid", start_color="D9E1F2")
SUB_FONT         = Font(name="Arial", bold=True, color="1F4E78", size=10)

# Legacy metrics group — medium blue
LEGACY_HDR_FILL  = PatternFill("solid", start_color="2E75B6")

# Migrated metrics group — green
MIGRATED_HDR_FILL = PatternFill("solid", start_color="548235")

# Comparison group — amber
COMPARE_HDR_FILL  = PatternFill("solid", start_color="BF8F00")

BODY_FONT        = Font(name="Arial", size=10)
BORDER           = _thin_border()
CENTER           = Alignment(horizontal="center", vertical="center")
LEFT             = Alignment(horizontal="left",   vertical="center")

# Verdict conditional format colours
VERDICT_COLOURS = {
    "Faster":             ("C6EFCE", "006100"),   # green fill, dark green text
    "Slower":             ("FFC7CE", "9C0006"),   # red fill, dark red text
    "Comparable":         ("FFEB9C", "9C5700"),   # yellow fill, amber text
    "Insufficient Runs":  ("D9D9D9", "595959"),   # grey fill, dark grey text (italic)
    "No Migrated Data":   ("F4B084", "833C0C"),   # orange fill, brown text (italic)
    "No Legacy Baseline": ("BDD7EE", "1F4E78"),   # light blue fill, dark blue text (italic)
}


# ─────────────────────────────────────────────────────────────────────────────
# Data collection
# ─────────────────────────────────────────────────────────────────────────────

def collect_triggers(wb, sheet_name: str) -> list:
    """Read unique (trigger, pipeline) pairs from a raw data sheet.

    Reads columns A (trigger) and B (pipeline) starting from row 2.
    Deduplicates while preserving first-seen order.

    Returns list of (trigger_name, pipeline_name) tuples.
    """
    ws   = wb[sheet_name]
    seen = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        trig, pipe = row[0], row[1]
        if trig and (trig, pipe) not in seen:
            seen[(trig, pipe)] = True
    return list(seen.keys())


def merge_and_sort_triggers(legacy: list, migrated: list) -> list:
    """Merge trigger lists from both environments, deduplicate, sort alphabetically."""
    combined = []
    seen     = set()
    for pair in legacy + migrated:
        if pair not in seen:
            seen.add(pair)
            combined.append(pair)
    combined.sort(key=lambda x: (x[0] or "").lower())
    return combined


# ─────────────────────────────────────────────────────────────────────────────
# Report builder
# ─────────────────────────────────────────────────────────────────────────────

def build_comparison_sheet(
    wb,
    all_triggers: list,
    min_runs: int,
    threshold: float,
) -> None:
    """Create (or overwrite) the Comparison sheet in the workbook.

    Layout (columns A–R, with E–H and K–P hidden):
      A  : Trigger Name
      B  : Pipeline Name
      C  : Legacy run count       (COUNTIFS formula)
      D  : Legacy avg duration    (AVERAGEIFS formula, seconds)
      I  : Migrated run count     (COUNTIFS formula)
      J  : Migrated avg duration  (AVERAGEIFS formula, seconds)
      Q  : Verdict                (nested IF formula)
      R  : Difference in minutes  (J - D) / 60

    Columns E–H and K–P are intentionally hidden to create visual
    separation between the Legacy, Migrated, and Comparison groups.
    """
    # Remove existing sheet if present
    if OUTPUT_SHEET in wb.sheetnames:
        del wb[OUTPUT_SHEET]
    ws = wb.create_sheet(OUTPUT_SHEET)

    # ── Row 1: Report title ───────────────────────────────────────────────
    ws["A1"] = "Trigger Performance Comparison: Legacy Environment vs Migrated Environment"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E78")
    ws.merge_cells("A1:R1")

    # ── Row 2: Group headers ──────────────────────────────────────────────
    _set(ws, 2, 1, "Trigger",              HEADER_FILL,       HEADER_FONT)
    _set(ws, 2, 2, "Pipeline",             HEADER_FILL,       HEADER_FONT)
    _set(ws, 2, 3, "Legacy Metrics",       LEGACY_HDR_FILL,   HEADER_FONT)
    ws.merge_cells("C2:D2")
    _set(ws, 2, 9, "Migrated Metrics",     MIGRATED_HDR_FILL, HEADER_FONT)
    ws.merge_cells("I2:J2")
    _set(ws, 2, 17, "Comparison",          COMPARE_HDR_FILL,  HEADER_FONT)
    ws.merge_cells("Q2:R2")

    for col in (1, 2, 3, 9, 17):
        ws.cell(row=2, column=col).alignment = CENTER
        ws.cell(row=2, column=col).border    = BORDER

    # ── Row 3: Sub-headers ────────────────────────────────────────────────
    sub_headers = {
        1:  "Trigger",
        2:  "Pipeline",
        3:  "Runs",
        4:  "Avg (s)",
        9:  "Runs",
        10: "Avg (s)",
        17: "Verdict",
        18: "Difference In Minutes",
    }
    for col, label in sub_headers.items():
        c           = ws.cell(row=3, column=col, value=label)
        c.fill      = SUB_FILL
        c.font      = SUB_FONT
        c.alignment = CENTER
        c.border    = BORDER

    # ── Data rows (row 4 onwards) ─────────────────────────────────────────
    for idx, (trig, pipe) in enumerate(all_triggers):
        r = 4 + idx

        ws.cell(row=r, column=1, value=trig)
        ws.cell(row=r, column=2, value=pipe or "")

        # Legacy: run count (C) and average duration (D)
        ws.cell(row=r, column=3).value = (
            f'=COUNTIFS({LEGACY_SHEET}!$A:$A,A{r},{LEGACY_SHEET}!$E:$E,"Succeeded")'
        )
        ws.cell(row=r, column=4).value = (
            f'=IFERROR(AVERAGEIFS({LEGACY_SHEET}!$F:$F,'
            f'{LEGACY_SHEET}!$A:$A,A{r},{LEGACY_SHEET}!$E:$E,"Succeeded"),"")'
        )

        # Migrated: run count (I) and average duration (J)
        ws.cell(row=r, column=9).value = (
            f'=COUNTIFS({MIGRATED_SHEET}!$A:$A,A{r},{MIGRATED_SHEET}!$E:$E,"Succeeded")'
        )
        ws.cell(row=r, column=10).value = (
            f'=IFERROR(AVERAGEIFS({MIGRATED_SHEET}!$F:$F,'
            f'{MIGRATED_SHEET}!$A:$A,A{r},{MIGRATED_SHEET}!$E:$E,"Succeeded"),"")'
        )

        # Verdict (Q): nested IF comparing migrated avg vs legacy avg
        ws.cell(row=r, column=17).value = (
            f'=IF(I{r}=0,"No Migrated Data",'
            f'IF(I{r}<{min_runs},"Insufficient Runs",'
            f'IF(C{r}=0,"No Legacy Baseline",'
            f'IF((J{r}-D{r})/D{r}<=-{threshold},"Faster",'
            f'IF((J{r}-D{r})/D{r}>={threshold},"Slower","Comparable")))))'
        )

        # Difference in minutes (R): (Migrated avg - Legacy avg) / 60
        ws.cell(row=r, column=18).value = (
            f'=IF(OR(D{r}="",J{r}=""),"",(J{r}-D{r})/60)'
        )

        # Apply styles to all columns in this row
        for col in range(1, 19):
            cell        = ws.cell(row=r, column=col)
            cell.font   = BODY_FONT
            cell.border = BORDER
            if col in (3, 9):
                cell.alignment     = CENTER
                cell.number_format = "0"
            elif col in (4, 10):
                cell.number_format = "#,##0;(#,##0);-"
                cell.alignment     = CENTER
            elif col == 18:
                cell.number_format = "#,##0.00;(#,##0.00);-"
                cell.alignment     = CENTER
            else:
                cell.alignment = LEFT

    last_row      = 3 + len(all_triggers)
    verdict_range = f"Q4:Q{last_row}"

    # ── Conditional formatting on Verdict column ───────────────────────────
    italic_statuses = {"Insufficient Runs", "No Migrated Data", "No Legacy Baseline"}
    for verdict, (fill_hex, font_hex) in VERDICT_COLOURS.items():
        ws.conditional_formatting.add(
            verdict_range,
            FormulaRule(
                formula=[f'EXACT(Q4,"{verdict}")'],
                fill=PatternFill("solid", start_color=fill_hex),
                font=Font(
                    color=font_hex,
                    bold=(verdict not in italic_statuses),
                    italic=(verdict in italic_statuses),
                ),
            ),
        )

    # ── Hide spacer columns (E–H, K–P) ────────────────────────────────────
    for col_letter in ["E", "F", "G", "H", "K", "L", "M", "N", "O", "P"]:
        ws.column_dimensions[col_letter].hidden = True

    # ── Column widths ──────────────────────────────────────────────────────
    col_widths = {
        "A": 32, "B": 38,
        "C": 9,  "D": 12,
        "I": 9,  "J": 12,
        "Q": 22, "R": 24,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # ── Freeze panes and row heights ──────────────────────────────────────
    ws.freeze_panes       = "C4"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 26


def _set(ws, row: int, col: int, value: str, fill: PatternFill, font: Font) -> None:
    """Helper: set value, fill, and font on a cell."""
    c       = ws.cell(row=row, column=col, value=value)
    c.fill  = fill
    c.font  = font


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description=(
            "Build a formatted ADF trigger performance comparison sheet "
            f"(Legacy Environment vs Migrated Environment) in an Excel workbook."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
Input workbook must contain:
  {LEGACY_SHEET}   — trigger run data from the legacy environment
  {MIGRATED_SHEET} — trigger run data from the migrated environment

Columns expected in each sheet:
  A: Trigger Name
  B: Pipeline Name
  E: Run Status (e.g. "Succeeded")
  F: Duration in seconds (numeric)

Examples:
  python adf_trigger_comparison.py
  python adf_trigger_comparison.py --input my_workbook.xlsx
  python adf_trigger_comparison.py --input data.xlsx --min-runs 3 --threshold 0.15
        """,
    )
    p.add_argument(
        "--input", "-i",
        default="ADF_Trigger_Raw.xlsx",
        help="Path to the input Excel workbook (default: ADF_Trigger_Raw.xlsx)",
    )
    p.add_argument(
        "--min-runs",
        type=int,
        default=DEFAULT_MIN_RUNS,
        help=f"Minimum migrated runs required for a valid verdict (default: {DEFAULT_MIN_RUNS})",
    )
    p.add_argument(
        "--threshold",
        type=float,
        default=DEFAULT_THRESHOLD,
        help=(
            f"Percentage difference threshold for Faster/Slower verdict "
            f"(default: {DEFAULT_THRESHOLD} = {int(DEFAULT_THRESHOLD * 100)}%%)"
        ),
    )
    return p.parse_args()


def main() -> None:
    args = parse_args()

    print(f"Loading workbook: {args.input}")
    try:
        wb = load_workbook(args.input)
    except FileNotFoundError:
        print(f"[ERROR] File not found: {args.input}")
        sys.exit(1)

    missing = [s for s in (LEGACY_SHEET, MIGRATED_SHEET) if s not in wb.sheetnames]
    if missing:
        print(f"[ERROR] Missing required sheets: {missing}")
        print(f"        Expected: '{LEGACY_SHEET}' and '{MIGRATED_SHEET}'")
        sys.exit(1)

    print(f"Reading trigger data from '{LEGACY_SHEET}' and '{MIGRATED_SHEET}'...")
    legacy_triggers   = collect_triggers(wb, LEGACY_SHEET)
    migrated_triggers = collect_triggers(wb, MIGRATED_SHEET)
    all_triggers      = merge_and_sort_triggers(legacy_triggers, migrated_triggers)

    print(f"  Legacy triggers   : {len(legacy_triggers)}")
    print(f"  Migrated triggers : {len(migrated_triggers)}")
    print(f"  Combined unique   : {len(all_triggers)}")

    print(f"\nBuilding '{OUTPUT_SHEET}' sheet...")
    print(f"  Min runs threshold : {args.min_runs}")
    print(f"  Verdict threshold  : {int(args.threshold * 100)}%")
    build_comparison_sheet(wb, all_triggers, args.min_runs, args.threshold)

    wb.save(args.input)
    print(f"\nDone. '{OUTPUT_SHEET}' sheet rebuilt with {len(all_triggers)} triggers.")
    print(f"Workbook saved: {args.input}")


if __name__ == "__main__":
    main()
