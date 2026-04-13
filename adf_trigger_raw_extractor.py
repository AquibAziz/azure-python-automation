#!/usr/bin/env python3
"""
ADF Trigger Run Raw Data Extractor
====================================
Extracts trigger run history from one or both ADF environments (Legacy and
Migrated) and writes the raw data into named sheets in an Excel workbook.

The output workbook is consumed by adf_trigger_comparison.py to build the
formatted performance comparison report.

Output sheets written:
  Legacy_Raw   — trigger run data from the legacy environment
  Migrated_Raw — trigger run data from the migrated environment

Output columns (both sheets):
  A: Trigger Name
  B: Pipeline Name
  C: Run Start Time
  D: Run End Time
  E: Status
  F: Duration (seconds)
  G: Duration (minutes)
  H: Run Date
  I: Environment label

How it works:
  1. Queries all trigger runs in the specified date window via ARM REST API
  2. For each trigger run, fetches the associated pipeline run details
  3. Validates invokedBy to confirm the trigger → pipeline relationship
  4. Writes all valid rows to the output Excel sheet

Authentication:
  Uses DefaultAzureCredential. Ensure one of the following:
    - az login  (Azure CLI)
    - Managed Identity  (when running in Azure)
    - Environment variables: AZURE_CLIENT_ID, AZURE_CLIENT_SECRET, AZURE_TENANT_ID

Configuration:
  All environment-specific values are loaded from environment variables.
  Never hardcode subscription IDs, resource group names, or factory names.

  Legacy environment:
    LEGACY_SUBSCRIPTION_ID
    LEGACY_RESOURCE_GROUP
    LEGACY_FACTORY_NAME

  Migrated environment:
    MIGRATED_SUBSCRIPTION_ID
    MIGRATED_RESOURCE_GROUP
    MIGRATED_FACTORY_NAME

Usage:
  # Extract both environments, last 45 days (default)
  python adf_trigger_raw_extractor.py

  # Extract last 30 days
  python adf_trigger_raw_extractor.py --days 30

  # Extract specific date range
  python adf_trigger_raw_extractor.py --start 2024-01-01 --end 2024-01-31

  # Extract legacy environment only
  python adf_trigger_raw_extractor.py --env legacy

  # Extract migrated environment only
  python adf_trigger_raw_extractor.py --env migrated

  # Custom output file and worker count
  python adf_trigger_raw_extractor.py --output my_data.xlsx --workers 4
"""

import os
import sys
import time
import argparse
import threading
from datetime import datetime, timedelta, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from azure.identity import DefaultAzureCredential
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

API_VERSION           = "2018-06-01"
MGMT_SCOPE            = "https://management.azure.com/.default"
REQUEST_TIMEOUT       = (10, 60)   # (connect timeout, read timeout) in seconds
DEFAULT_WORKERS       = 2          # conservative default — increase with caution
TOKEN_REFRESH_SECONDS = 2400       # refresh token every 40 min (tokens last ~60 min)
MAX_RETRIES           = 4

# Environment labels used in the Excel sheet names and row data
LEGACY_LABEL   = "Legacy"
MIGRATED_LABEL = "Migrated"
LEGACY_SHEET   = "Legacy_Raw"
MIGRATED_SHEET = "Migrated_Raw"


# ─────────────────────────────────────────────────────────────────────────────
# Thread-safe token cache
# ─────────────────────────────────────────────────────────────────────────────

_thread_local = threading.local()   # per-thread HTTP session
_token_lock   = threading.Lock()
_token_cache  = {"token": None, "expires_at": 0}
_credential   = None


def get_credential() -> DefaultAzureCredential:
    """Return a singleton DefaultAzureCredential instance."""
    global _credential
    if _credential is None:
        _credential = DefaultAzureCredential()
    return _credential


def get_fresh_token(force: bool = False) -> str:
    """Return a valid Azure ARM bearer token.

    Thread-safe. Refreshes proactively when within TOKEN_REFRESH_SECONDS
    of expiry, or immediately when force=True (e.g. after a 401 response).
    """
    with _token_lock:
        now = time.time()
        needs_refresh = (
            force
            or _token_cache["token"] is None
            or now >= _token_cache["expires_at"]
        )
        if needs_refresh:
            reason = "forced refresh" if force else "token expiry"
            print(f"  [token] refreshing Azure AD token ({reason})...")
            token_obj = get_credential().get_token(MGMT_SCOPE)
            _token_cache["token"]      = token_obj.token
            _token_cache["expires_at"] = now + TOKEN_REFRESH_SECONDS
            expiry_str = datetime.fromtimestamp(_token_cache["expires_at"]).strftime(
                "%Y-%m-%d %H:%M:%S"
            )
            print(f"  [token] cached — valid until {expiry_str}")
        return _token_cache["token"]


# ─────────────────────────────────────────────────────────────────────────────
# HTTP session (per-thread connection pooling)
# ─────────────────────────────────────────────────────────────────────────────

def get_session() -> requests.Session:
    """Return a thread-local requests.Session with connection pooling."""
    if not hasattr(_thread_local, "session"):
        s       = requests.Session()
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=50, pool_maxsize=50, max_retries=3
        )
        s.mount("https://", adapter)
        _thread_local.session = s
    return _thread_local.session


# ─────────────────────────────────────────────────────────────────────────────
# ARM API helpers
# ─────────────────────────────────────────────────────────────────────────────

def _iso(dt: datetime) -> str:
    """Format a datetime as an ISO 8601 string with Z suffix."""
    return dt.isoformat().replace("+00:00", "Z")


def _retry_request(func, max_retries: int = MAX_RETRIES):
    """Execute an HTTP request function with retry/backoff logic.

    Handles:
      401 (auth expired)  → force-refresh token and retry
      429 (throttled)     → respect Retry-After header
      5xx (server error)  → exponential backoff
      ConnectionError     → exponential backoff
    """
    for attempt in range(max_retries):
        try:
            r = func()
        except requests.exceptions.RequestException:
            if attempt == max_retries - 1:
                raise
            time.sleep(2 ** attempt)
            continue

        if r.status_code == 401:
            get_fresh_token(force=True)
            if attempt == max_retries - 1:
                r.raise_for_status()
            continue

        if r.status_code == 429 or r.status_code >= 500:
            retry_after = r.headers.get("Retry-After")
            try:
                wait = int(retry_after) if retry_after else 2 ** (attempt + 1)
            except (TypeError, ValueError):
                wait = 2 ** (attempt + 1)
            if attempt == max_retries - 1:
                r.raise_for_status()
            time.sleep(wait)
            continue

        r.raise_for_status()
        return r

    raise RuntimeError("Max retries exceeded")


def query_trigger_runs(base_url: str, start: datetime, end: datetime) -> list:
    """Query all trigger runs in the given time window, handling pagination.

    Uses POST /queryTriggerRuns with continuationToken pagination.
    Prints progress per page to stderr for visibility during long windows.
    """
    url     = f"{base_url}/queryTriggerRuns?api-version={API_VERSION}"
    body    = {"lastUpdatedAfter": _iso(start), "lastUpdatedBefore": _iso(end)}
    session = get_session()
    results = []
    page    = 0

    while True:
        page += 1

        def do_post():
            headers = {
                "Authorization": f"Bearer {get_fresh_token()}",
                "Content-Type":  "application/json",
            }
            return session.post(url, json=body, headers=headers, timeout=REQUEST_TIMEOUT)

        r    = _retry_request(do_post)
        data = r.json()
        batch = data.get("value", [])
        results.extend(batch)
        print(f"  trigger runs page {page}: +{len(batch)} (total {len(results)})")

        cont = data.get("continuationToken")
        if not cont:
            break
        body["continuationToken"] = cont

    return results


def get_pipeline_run(base_url: str, run_id: str) -> dict | None:
    """Fetch a single pipeline run by run ID.

    Returns None if the run is not found (404).
    """
    url     = f"{base_url}/pipelineruns/{run_id}?api-version={API_VERSION}"
    session = get_session()

    def do_get():
        headers = {"Authorization": f"Bearer {get_fresh_token()}"}
        return session.get(url, headers=headers, timeout=REQUEST_TIMEOUT)

    try:
        r = _retry_request(do_get)
        return r.json()
    except requests.exceptions.HTTPError as exc:
        if exc.response is not None and exc.response.status_code == 404:
            return None
        raise


# ─────────────────────────────────────────────────────────────────────────────
# Trigger run processor
# ─────────────────────────────────────────────────────────────────────────────

def process_trigger_run(tr: dict, base_url: str) -> tuple[list, dict]:
    """Process a single trigger run and return valid data rows.

    For each pipeline run ID referenced in the trigger run:
      1. Fetch the pipeline run details
      2. Validate that invokedBy matches the trigger name
      3. Parse timestamps and calculate duration
      4. Return a data row

    Designed to be called from worker threads — reads only thread-safe globals.

    Returns:
        rows  : list of data rows (each row is a list of column values)
        drops : dict of drop/error reason counts for diagnostics
    """
    rows  = []
    drops = {}

    try:
        trigger_name = tr.get("triggerName")
        if not trigger_name:
            drops["missing_trigger_name"] = 1
            return rows, drops

        triggered_pipelines = tr.get("triggeredPipelines") or {}
        if not triggered_pipelines:
            drops["no_triggered_pipelines"] = 1
            return rows, drops

        for pipeline_run_id in triggered_pipelines.values():
            if not pipeline_run_id:
                drops["empty_pipeline_run_id"] = drops.get("empty_pipeline_run_id", 0) + 1
                continue

            pr = get_pipeline_run(base_url, pipeline_run_id)
            if not pr:
                drops["pipeline_run_not_found"] = drops.get("pipeline_run_not_found", 0) + 1
                continue

            # Validate invokedBy — skip if mismatched (avoids cross-trigger contamination)
            invoked_by   = pr.get("invokedBy") or {}
            invoker_name = invoked_by.get("name", "")
            if invoker_name and invoker_name != trigger_name:
                drops["invokedby_mismatch"] = drops.get("invokedby_mismatch", 0) + 1
                continue

            pipeline_name = pr.get("pipelineName")
            run_start     = pr.get("runStart")
            run_end       = pr.get("runEnd")
            status        = pr.get("status")

            if not run_start or not run_end:
                drops["missing_timestamps"] = drops.get("missing_timestamps", 0) + 1
                continue

            try:
                rs  = datetime.fromisoformat(run_start.replace("Z", "+00:00")).replace(tzinfo=None)
                re_ = datetime.fromisoformat(run_end.replace("Z",   "+00:00")).replace(tzinfo=None)
            except (ValueError, AttributeError):
                drops["timestamp_parse_failed"] = drops.get("timestamp_parse_failed", 0) + 1
                continue

            duration_sec = (re_ - rs).total_seconds()
            rows.append([
                trigger_name,
                pipeline_name,
                rs,
                re_,
                status,
                duration_sec,
                duration_sec / 60.0,
                rs.date(),
            ])

    except Exception as exc:
        err_key         = type(exc).__name__
        drops[err_key]  = drops.get(err_key, 0) + 1

    return rows, drops


# ─────────────────────────────────────────────────────────────────────────────
# Extraction orchestrator
# ─────────────────────────────────────────────────────────────────────────────

def extract(
    env_label: str,
    subscription_id: str,
    resource_group: str,
    factory_name: str,
    start: datetime,
    end: datetime,
    workers: int,
) -> list:
    """Orchestrate the full extraction for one environment.

    Steps:
      1. Query all trigger runs in the time window (paginated)
      2. Fetch pipeline run details in parallel (thread pool)
      3. Merge results and report diagnostics

    Returns list of data rows, each including the env_label as final element.
    """
    base_url = (
        f"https://management.azure.com"
        f"/subscriptions/{subscription_id}"
        f"/resourceGroups/{resource_group}"
        f"/providers/Microsoft.DataFactory/factories/{factory_name}"
    )

    print(f"\n{'=' * 60}")
    print(f"[{env_label}] Starting extraction")
    print(f"[{env_label}] Factory       : {factory_name}")
    print(f"[{env_label}] Resource Group: {resource_group}")
    print(f"[{env_label}] Window        : {_iso(start)} → {_iso(end)}")
    print(f"[{env_label}] Workers       : {workers}")
    print(f"{'=' * 60}")

    print(f"[{env_label}] Step 1: querying trigger runs...")
    trigger_runs = query_trigger_runs(base_url, start, end)
    print(f"[{env_label}] Trigger runs found: {len(trigger_runs)}")

    if not trigger_runs:
        print(f"[{env_label}] No trigger runs in window — nothing to extract.")
        return []

    print(f"[{env_label}] Step 2: fetching pipeline run details ({workers} workers)...")
    all_rows      = []
    error_summary = {}
    completed     = 0
    t0            = time.time()
    progress_lock = threading.Lock()

    def merge_drops(drops: dict) -> None:
        for k, v in drops.items():
            error_summary[k] = error_summary.get(k, 0) + v

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {
            executor.submit(process_trigger_run, tr, base_url): i
            for i, tr in enumerate(trigger_runs)
        }
        for future in as_completed(futures):
            try:
                rows, drops = future.result()
                with progress_lock:
                    for row in rows:
                        all_rows.append(row + [env_label])
                    merge_drops(drops)
                    completed += 1

                    if completed % 50 == 0 or completed == len(trigger_runs):
                        elapsed = time.time() - t0
                        rate    = completed / max(elapsed, 0.1)
                        eta     = (len(trigger_runs) - completed) / max(rate, 0.1)
                        print(
                            f"  [{env_label}] {completed}/{len(trigger_runs)} "
                            f"({rate:.1f}/sec, ETA {round(eta)}s, rows={len(all_rows)})"
                        )
            except Exception as exc:
                with progress_lock:
                    key                  = type(exc).__name__
                    error_summary[key]   = error_summary.get(key, 0) + 1

    total_time = time.time() - t0
    print(f"[{env_label}] Done — {len(all_rows)} rows in {total_time:.1f}s")
    if error_summary:
        print(f"[{env_label}] Drop/error breakdown:")
        for reason, count in sorted(error_summary.items(), key=lambda x: -x[1]):
            print(f"  {reason}: {count}")

    return all_rows


# ─────────────────────────────────────────────────────────────────────────────
# Excel writer
# ─────────────────────────────────────────────────────────────────────────────

def write_sheet(wb_path: str, sheet_name: str, rows: list) -> None:
    """Write extracted rows to a named sheet in the output workbook.

    Creates the workbook if it doesn't exist.
    Replaces the sheet if it already exists.
    """
    print(f"Writing {len(rows):,} rows to sheet '{sheet_name}'...")

    if os.path.exists(wb_path):
        wb = load_workbook(wb_path)
    else:
        wb = Workbook()
        # Remove default empty sheet created by openpyxl
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            del wb["Sheet"]

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Header row
    headers     = [
        "Trigger Name", "Pipeline Name",
        "Run Start Time", "Run End Time",
        "Status",
        "Duration (s)", "Duration (m)",
        "Run Date", "Environment",
    ]
    header_fill = PatternFill("solid", start_color="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row[0])
        ws.cell(row=row_idx, column=2, value=row[1])
        ws.cell(row=row_idx, column=3, value=row[2]).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row=row_idx, column=4, value=row[3]).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row=row_idx, column=5, value=row[4])
        ws.cell(row=row_idx, column=6, value=row[5]).number_format = "#,##0"
        ws.cell(row=row_idx, column=7, value=row[6]).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=8, value=row[7]).number_format = "yyyy-mm-dd"
        ws.cell(row=row_idx, column=9, value=row[8])

    # Column widths
    for col, width in enumerate([28, 32, 20, 20, 12, 12, 12, 13, 13], 1):
        ws.column_dimensions[chr(64 + col)].width = width
    ws.freeze_panes = "A2"

    wb.save(wb_path)
    print(f"Saved: {wb_path}")


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def _require_env(var: str) -> str:
    """Read a required environment variable or exit with a clear error."""
    val = os.environ.get(var)
    if not val:
        print(f"[ERROR] Required environment variable not set: {var}")
        print(f"        Add it to your .env file or export it before running.")
        sys.exit(1)
    return val


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Extract ADF trigger run data for Legacy and/or Migrated environments.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Required environment variables:
  Legacy environment:
    LEGACY_SUBSCRIPTION_ID
    LEGACY_RESOURCE_GROUP
    LEGACY_FACTORY_NAME

  Migrated environment:
    MIGRATED_SUBSCRIPTION_ID
    MIGRATED_RESOURCE_GROUP
    MIGRATED_FACTORY_NAME

Examples:
  # Both environments, last 45 days
  python adf_trigger_raw_extractor.py

  # Last 30 days
  python adf_trigger_raw_extractor.py --days 30

  # Specific date range
  python adf_trigger_raw_extractor.py --start 2024-01-01 --end 2024-01-31

  # Legacy only
  python adf_trigger_raw_extractor.py --env legacy

  # Migrated only
  python adf_trigger_raw_extractor.py --env migrated

  # Custom output file and parallelism
  python adf_trigger_raw_extractor.py --output my_data.xlsx --workers 4
        """,
    )
    p.add_argument("--start",   help="Start date (YYYY-MM-DD)")
    p.add_argument("--end",     help="End date (YYYY-MM-DD)")
    p.add_argument("--days",    type=int, help="Days back from today (default: 45)")
    p.add_argument("--env",     choices=["legacy", "migrated", "both"], default="both",
                   help="Which environment to extract (default: both)")
    p.add_argument("--output",  default="ADF_Trigger_Raw.xlsx",
                   help="Output Excel file path (default: ADF_Trigger_Raw.xlsx)")
    p.add_argument("--workers", type=int, default=DEFAULT_WORKERS,
                   help=f"Parallel worker threads (default: {DEFAULT_WORKERS})")
    return p.parse_args()


def main() -> None:
    args = parse_args()

    # Resolve time window
    end = datetime.now(timezone.utc)
    if args.end:
        end = datetime.fromisoformat(args.end + "T23:59:59").replace(tzinfo=timezone.utc)

    if args.start:
        start = datetime.fromisoformat(args.start + "T00:00:00").replace(tzinfo=timezone.utc)
    elif args.days:
        start = end - timedelta(days=args.days)
    else:
        start = end - timedelta(days=45)

    print(f"Extraction window: {_iso(start)} → {_iso(end)}")
    print(f"Output file      : {args.output}")
    print(f"Environments     : {args.env}")

    # Prime the token cache once at startup so any auth failures are visible immediately
    get_fresh_token()

    if args.env in ("legacy", "both"):
        rows = extract(
            env_label       = LEGACY_LABEL,
            subscription_id = _require_env("LEGACY_SUBSCRIPTION_ID"),
            resource_group  = _require_env("LEGACY_RESOURCE_GROUP"),
            factory_name    = _require_env("LEGACY_FACTORY_NAME"),
            start=start, end=end, workers=args.workers,
        )
        write_sheet(args.output, LEGACY_SHEET, rows)

    if args.env in ("migrated", "both"):
        rows = extract(
            env_label       = MIGRATED_LABEL,
            subscription_id = _require_env("MIGRATED_SUBSCRIPTION_ID"),
            resource_group  = _require_env("MIGRATED_RESOURCE_GROUP"),
            factory_name    = _require_env("MIGRATED_FACTORY_NAME"),
            start=start, end=end, workers=args.workers,
        )
        write_sheet(args.output, MIGRATED_SHEET, rows)

    print(f"\nAll done. Open {args.output} to view raw data.")
    print(f"Run adf_trigger_comparison.py --input {args.output} to build the comparison sheet.")


if __name__ == "__main__":
    main()
